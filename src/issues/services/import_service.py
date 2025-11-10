"""
Issue import services.
"""

import csv
from io import StringIO, BytesIO
from typing import List, Dict, Tuple
from django.contrib.auth.models import User
from issues.models import Issue, Project, Customer


def parse_csv_file(file_content: bytes) -> List[Dict]:
    """解析 CSV 檔案"""
    content = file_content.decode('utf-8-sig')  # 處理 BOM
    reader = csv.DictReader(StringIO(content))
    return list(reader)


def parse_xlsx_file(file_content: bytes) -> List[Dict]:
    """解析 XLSX 檔案"""
    from openpyxl import load_workbook
    
    wb = load_workbook(BytesIO(file_content), data_only=True)
    ws = wb.active
    
    # 讀取標題行
    headers = []
    for cell in ws[1]:
        headers.append(cell.value or '')
    
    # 讀取資料行
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # 跳過空行
            continue
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = value
        if row_dict:
            rows.append(row_dict)
    
    return rows


def import_issues(data: List[Dict], default_user: User = None) -> Tuple[int, int, List[str]]:
    """
    匯入 Issue
    
    Returns:
        (成功數量, 失敗數量, 錯誤訊息列表)
    """
    success_count = 0
    error_count = 0
    errors = []
    
    # 欄位對應（Excel 欄位名稱 -> 模型欄位）
    field_mapping = {
        'ID': None,  # 忽略，自動生成
        '標題': 'title',
        '描述': 'description',
        '狀態': 'status',
        '優先級': 'priority',
        '類別': 'category',
        '來源': 'source',
        '專案': 'project',
        '客戶': 'customer',
        '負責人': 'assignee',
        '回報人': 'reporter',
    }
    
    for idx, row in enumerate(data, start=2):  # 從第 2 行開始（第 1 行是標題）
        try:
            issue_data = {}
            
            # 處理基本欄位
            for excel_field, model_field in field_mapping.items():
                if not model_field:
                    continue
                
                value = row.get(excel_field, '').strip() if row.get(excel_field) else ''
                
                if model_field in ['title', 'description', 'status', 'priority', 'category', 'source']:
                    issue_data[model_field] = value
                elif model_field == 'project':
                    if value:
                        project, _ = Project.objects.get_or_create(name=value, defaults={'code': value[:10]})
                        issue_data['project'] = project
                elif model_field == 'customer':
                    if value:
                        customer, _ = Customer.objects.get_or_create(name=value, defaults={'code': value[:10]})
                        issue_data['customer'] = customer
                elif model_field in ['assignee', 'reporter']:
                    if value:
                        try:
                            user = User.objects.get(username=value)
                            issue_data[model_field] = user
                        except User.DoesNotExist:
                            # 如果用戶不存在，使用預設用戶
                            if default_user:
                                issue_data[model_field] = default_user
            
            # 驗證必填欄位
            if not issue_data.get('title'):
                errors.append(f'第 {idx} 行：標題為必填')
                error_count += 1
                continue
            
            if not issue_data.get('description'):
                issue_data['description'] = issue_data.get('title', '')
            
            # 設定預設值
            if not issue_data.get('status'):
                issue_data['status'] = 'Open'
            if not issue_data.get('priority'):
                issue_data['priority'] = 'Medium'
            if not issue_data.get('category'):
                issue_data['category'] = '其他'
            if not issue_data.get('source'):
                issue_data['source'] = '業務回報'
            
            # 設定 reporter（如果沒有）
            if not issue_data.get('reporter') and default_user:
                issue_data['reporter'] = default_user
            
            # 建立 Issue
            Issue.objects.create(**issue_data)
            success_count += 1
            
        except Exception as e:
            errors.append(f'第 {idx} 行：{str(e)}')
            error_count += 1
    
    return success_count, error_count, errors

